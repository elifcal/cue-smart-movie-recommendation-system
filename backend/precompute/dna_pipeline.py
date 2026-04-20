import os
import time
import random
import argparse
import whisper
import sys
import yt_dlp
from supabase import create_client
from dotenv import load_dotenv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Fragman sayılacak maksimum süre (saniye). Bunun üstündekiler full film modunda işlenir.
FRAGMAN_LIMIT_SANIYE = 15 * 60  # 15 dakika

# Full film modunda indirilecek maksimum süre (saniye)
KISIM_SURE_SANIYE = 120  # ilk 2 dakika

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from analysis.audio import ses_indir, analyze_audio
from analysis.visual import analyze_visual

load_dotenv()
supabase = create_client(os.getenv("SUPABASE_URL"), os.getenv("SUPABASE_KEY"))

print("🧠 Whisper (Audio DNA) modeli yükleniyor...")
model = whisper.load_model("base")

hatali_filmler = []


def zaten_islendi(tmdb_id):
    r = supabase.table("film_dna").select("tmdb_id").eq("tmdb_id", tmdb_id).execute()
    return len(r.data) > 0


def video_sure_kontrol(youtube_key):
    """İndirmeden önce yt-dlp ile video süresini saniye cinsinden döndürür."""
    url = f"https://www.youtube.com/watch?v={youtube_key}"
    cookie_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cookies.txt")
    ydl_opts = {
        "quiet": True,
        "skip_download": True,
        "no_warnings": True,
        "cookiefile": cookie_path if os.path.exists(cookie_path) else None,
        "nocheckcertificate": True,
        "user_agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
        "remote_components": "ejs:github",
    }
    try:
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            info = ydl.extract_info(url, download=False)
            return info.get("duration", 0)
    except Exception as e:
        print(f"⚠️  Süre alınamadı ({youtube_key}): {e}")
        return None


def ses_indir_kisim(youtube_key, sure_saniye=120):
    """
    Full film gibi uzun videolardan sadece ilk `sure_saniye` saniyeyi indirir.
    yt-dlp stdout'a yazar, ffmpeg pipe üzerinden stdin'den okur ve -t ile durur.
    Tam indirme yapmaz, disk ve RAM dostu.
    """
    import subprocess

    kesik_yol = f"/tmp/{youtube_key}_kisim.mp3"

    if os.path.exists(kesik_yol):
        os.remove(kesik_yol)

    try:
        ytdlp_proc = subprocess.Popen(
            [
                "yt-dlp",
                "-f", "bestaudio/best",
                "-o", "-",
                "--quiet",
                "--no-warnings",
                f"https://www.youtube.com/watch?v={youtube_key}",
            ],
            stdout=subprocess.PIPE,
            stderr=subprocess.DEVNULL,
        )

        ffmpeg_proc = subprocess.Popen(
            [
                "ffmpeg", "-y",
                "-i", "pipe:0",
                "-t", str(sure_saniye),
                "-vn",
                "-acodec", "libmp3lame",
                "-ab", "128k",
                "-ar", "44100",
                kesik_yol,
            ],
            stdin=ytdlp_proc.stdout,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )

        ffmpeg_proc.wait(timeout=90)
        ytdlp_proc.terminate()

    except subprocess.TimeoutExpired:
        print(f"⚠️  Zaman aşımı: {youtube_key}")
        ytdlp_proc.kill()
        ffmpeg_proc.kill()
        return None
    except Exception as e:
        print(f"⚠️  Pipe hatası: {e}")
        return None

    return kesik_yol if os.path.exists(kesik_yol) else None


def hata_raporu_olustur(hatali_filmler, toplam_is):
    if not hatali_filmler:
        print("\n✅ Hata raporu oluşturulmadı — hiç hata yok!")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Hatalı Filmler"

    HEADER_FG = "FFFFFF"
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:E1")
    ws["A1"] = f"🎬 Film DNA — Hata Raporu  |  {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color=HEADER_FG)
    ws["A1"].fill = PatternFill("solid", fgColor="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:E2")
    ws["A2"] = (
        f"Toplam İşlenen: {toplam_is}  |  "
        f"Başarılı: {toplam_is - len(hatali_filmler)}  |  "
        f"Hatalı: {len(hatali_filmler)}"
    )
    ws["A2"].font = Font(name="Arial", italic=True, size=10, color="444444")
    ws["A2"].fill = PatternFill("solid", fgColor="DCE4FF")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    headers = ["#", "Film Adı (Title)", "TMDB ID", "YouTube URL", "Hata Mesajı"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, size=11, color=HEADER_FG)
        cell.fill = PatternFill("solid", fgColor="2E4DA3")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[3].height = 20

    for i, kayit in enumerate(hatali_filmler, 1):
        row = i + 3
        bg = "FFF0F0" if i % 2 == 0 else "FFFFFF"
        fill = PatternFill("solid", fgColor=bg)
        data = [
            i,
            kayit.get("title", "—"),
            kayit.get("tmdb_id", "—"),
            kayit.get("url", "—"),
            kayit.get("hata", "—"),
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 40

    col_widths = [5, 35, 12, 50, 60]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:E{3 + len(hatali_filmler)}"

    zaman_damgasi = datetime.now().strftime("%Y%m%d_%H%M%S")
    dosya_adi = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        f"hata_raporu_{zaman_damgasi}.xlsx"
    )
    wb.save(dosya_adi)
    print(f"\n📋 Hata raporu kaydedildi → {dosya_adi}  ({len(hatali_filmler)} hatalı film)")


def dna_isle(film, force=False):
    tmdb_id = film.get("tmdb_id")
    title = film.get("title")
    videos = film.get("videos", [])

    # Video key seçimi: Trailer > Teaser > İlk video
    youtube_key = None
    for v in videos:
        if v.get("type") == "Trailer" and v.get("site") == "YouTube":
            youtube_key = v.get("key")
            break
    if not youtube_key:
        for v in videos:
            if v.get("type") == "Teaser" and v.get("site") == "YouTube":
                youtube_key = v.get("key")
                break
    if not youtube_key and videos:
        youtube_key = videos[0].get("key")

    youtube_url = f"https://www.youtube.com/watch?v={youtube_key}" if youtube_key else "—"

    if not youtube_key:
        print(f"⚠️  Video bulunamadı: {title}")
        hatali_filmler.append({
            "title": title,
            "tmdb_id": tmdb_id,
            "url": "—",
            "hata": "YouTube video/key bulunamadı",
        })
        return

    if not force and zaten_islendi(tmdb_id):
        print(f"⏭  ATLANDI (zaten işlendi): {title}")
        print(f"   💡 Yeniden işlemek için --force ekle")
        return

    if force and zaten_islendi(tmdb_id):
        print(f"🔄 ZORLA YENİDEN İŞLENİYOR: {title}")
        supabase.table("film_dna").delete().eq("tmdb_id", tmdb_id).execute()

    print(f"🔍 Süre kontrol ediliyor: {title}")
    sure = video_sure_kontrol(youtube_key)

    if sure is None:
        print(f"⚠️  Süre alınamadı, atlanıyor: {title}")
        hatali_filmler.append({
            "title": title,
            "tmdb_id": tmdb_id,
            "url": youtube_url,
            "hata": "Video süresi metadata'dan alınamadı",
        })
        return

    full_film_modu = sure > FRAGMAN_LIMIT_SANIYE
    if full_film_modu:
        dk = sure // 60
        sn = sure % 60
        print(f"🎞️  FULL FİLM MODU ({dk} dk {sn} sn) — ilk {KISIM_SURE_SANIYE} sn alınacak: {title}")
    else:
        print(f"\n🎬 --- {title} İşleniyor --- ({sure // 60} dk {sure % 60} sn)")

    mp3_yolu = None
    try:
        if full_film_modu:
            mp3_yolu = ses_indir_kisim(youtube_key, sure_saniye=KISIM_SURE_SANIYE)
        else:
            mp3_yolu = ses_indir(youtube_key)

        if not mp3_yolu:
            hatali_filmler.append({
                "title": title,
                "tmdb_id": tmdb_id,
                "url": youtube_url,
                "hata": "Ses indirilemedi (video mevcut değil veya erişim engeli)",
            })
            return

        film_dili = film.get("original_language")
        ses_sonuclari = analyze_audio(mp3_yolu, model, sure=60, language=film_dili)
        poster_url = f"https://image.tmdb.org/t/p/w500{film.get('poster_path')}"
        gorsel_sonuclari = analyze_visual(tmdb_id, poster_url)

        supabase.table("film_dna").insert({
            "tmdb_id": tmdb_id,
            "title": title,
            "whisper_text": ses_sonuclari.get("text"),
            "emotion_curve": ses_sonuclari.get("emotion_curve"),
            "tempo": ses_sonuclari.get("tempo"),
            "energy": ses_sonuclari.get("energy"),
            "speech_ratio": ses_sonuclari.get("speech_ratio"),
            "color_palette": gorsel_sonuclari.get("color_palette"),
            "brightness": gorsel_sonuclari.get("brightness"),
            "warmth": gorsel_sonuclari.get("warmth"),
            "saturation": gorsel_sonuclari.get("saturation"),
        }).execute()
        print(f"✅ BAŞARILI{'  [full film modunda]' if full_film_modu else ''}: {title}")

        # Dil uyuşmazlığı varsa başarılı olsa da raporla
        if ses_sonuclari.get("dil_uyumsuzlugu"):
            hatali_filmler.append({
                "title": title,
                "tmdb_id": tmdb_id,
                "url": youtube_url,
                "hata": ses_sonuclari["dil_uyumsuzlugu"],
            })

    except Exception as e:
        print(f"⚠️  HATA: {e}")
        hatali_filmler.append({
            "title": title,
            "tmdb_id": tmdb_id,
            "url": youtube_url,
            "hata": str(e),
        })
    finally:
        if mp3_yolu and os.path.exists(mp3_yolu):
            os.remove(mp3_yolu)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Film DNA Pipeline")
    parser.add_argument("--offset",  type=int, default=0,    help="Kaçıncı filmden başla")
    parser.add_argument("--limit",   type=int, default=10,   help="Kaç film işle")
    parser.add_argument("--tmdb_id", type=int, default=None, help="Tek film: tmdb_id ile seç")
    parser.add_argument("--title",   type=str, default=None, help="Tek film: title ile ara (kısmi)")
    parser.add_argument("--force",   action="store_true",    help="Zaten işlenmiş filmleri yeniden işle")
    args = parser.parse_args()

    filmler = []

    if args.tmdb_id:
        res = supabase.table("movies").select("*").eq("tmdb_id", args.tmdb_id).execute()
        filmler = res.data
        if not filmler:
            print(f"❌ tmdb_id={args.tmdb_id} olan film bulunamadı.")
        else:
            print(f"🎯 Bulundu: {filmler[0].get('title')} (tmdb_id={args.tmdb_id})")

    elif args.title:
        res = supabase.table("movies").select("*").ilike("title", f"%{args.title}%").execute()
        filmler = res.data
        if not filmler:
            print(f"❌ '{args.title}' ile eşleşen film bulunamadı.")
        elif len(filmler) > 1:
            print(f"🔎 {len(filmler)} eşleşme bulundu:")
            for f in filmler:
                print(f"   tmdb_id={f.get('tmdb_id')}  |  {f.get('title')}")
            print("💡 Doğru filmi seçmek için --tmdb_id kullan.")
            filmler = []
        else:
            print(f"🎯 Bulundu: {filmler[0].get('title')} (tmdb_id={filmler[0].get('tmdb_id')})")

    else:
        res = (
            supabase.table("movies")
            .select("*")
            .order("tmdb_id")
            .range(args.offset, args.offset + args.limit - 1)
            .execute()
        )
        filmler = res.data

    if filmler:
        toplam_is = len(filmler)
        baslangic_zamani = time.time()

        for index, film in enumerate(filmler):
            print(f"\n📊 İlerleme: {index + 1} / {toplam_is}")
            dna_isle(film, force=args.force)
            if index < toplam_is - 1:
                time.sleep(random.randint(1, 3))

        toplam_sure_sn = time.time() - baslangic_zamani
        print(f"\n⏱️  Toplam Süre: {int(toplam_sure_sn // 60)} dk {int(toplam_sure_sn % 60)} sn")
        print(f"📈 Ortalama: {round(toplam_sure_sn / toplam_is, 2)} sn/film")

        hata_raporu_olustur(hatali_filmler, toplam_is)
    else:
        print("📭 İşlenecek film bulunamadı.")
